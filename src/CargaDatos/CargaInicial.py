from openpyxl import Workbook, load_workbook
import os
import sqlite3
from sqlite3 import Error
# from passlib.hash import sha512_crypt as sha512

def crearConeccion():

    db = None
    sqlitePath = os.getcwd() + '/data/PalmerasERP.db'

    try:
        db = sqlite3.connect(sqlitePath)
    except Error as e:
        print(f'No se pudo establecer la conección a la base de datos. Error ({e})')
    return db

def getPersonas():
    db = crearConeccion()
    cursor = db.cursor()
    selStr = '''SELECT * FROM ig_personas'''
    cursor.execute(selStr)
    personas = cursor.fetchall()
    db.close()
    return personas

def crearIndiceIdentificacion(personas):
    dictPersonas = { persona[4]: persona for persona in personas}
    return dictPersonas

def getLocales():
    db = crearConeccion()
    cursor = db.cursor()
    selStr = '''SELECT * FROM ig_Locales'''
    cursor.execute(selStr)
    locales = cursor.fetchall()
    db.close()
    return locales

def crearIndiceCodLocal(locales):
    dictLocales = { local[1]: local for local in locales}
    return dictLocales

def getPerfiles():
    db = crearConeccion()
    cursor = db.cursor()
    selStr = '''SELECT * FROM ig_perfiles'''
    cursor.execute(selStr)
    perfiles = cursor.fetchall()
    db.close()
    return perfiles

def crearIndicePerfil(perfiles):
    dictPerfiles = { perfil[1]: perfil for perfil in perfiles}
    return dictPerfiles

def getGastosGenerales():
    db = crearConeccion()
    cursor = db.cursor()
    selStr = '''SELECT * FROM loc_gastosGenerales'''
    cursor.execute(selStr)
    gastoGeneral = cursor.fetchall()
    db.close()
    return gastoGeneral

def crearIndiceGastosGenerales(gastosGenerales):
    dictEgresos = {egreso[1]: egreso for egreso in gastosGenerales}
    return dictEgresos

def seleccionarDatos(ws):
    columnas = [cell.value for cell in ws[10]]
    try:
        columnasConDatos = columnas.index('Verif')
    except:
        columnasConDatos = None
    dataRows = []
    for row in ws.iter_rows(min_row=11,values_only=True):
        # Convertir los campos None en campos vacíos
        listRow = list(row)
        for i in range(len(row)):
            if listRow[i] is None:
                listRow[i] = ''
        row=tuple(listRow)

        # Si hay una fila vacía terminar de buscar datos
        filaVacia = True
        for campo in row[0:columnasConDatos]:
            if campo != '':
                filaVacia = False
                break

        if filaVacia:
            break

        # Si hay datos después de la columna "Verif", eliminarlos
        if columnasConDatos:
            dataRows.append(row[0:columnasConDatos])
        else:
            dataRows.append(row)
    return dataRows


def carga_igLocales(ws):
    datos = seleccionarDatos(ws)
    db = crearConeccion()
    cursor = db.cursor()
    insStr = '''INSERT OR REPLACE INTO ig_locales (
                            codLocal,
                            local,
                            alias,
                            tipo,
                            ciudad,
                            direccion,
                            referencia,
                            codigoArea,
                            telefono,
                            codigoPostal,
                            estado)
                        VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)'''
    cursor.executemany(insStr, datos)
    db.commit()
    db.close()

def carga_igPersonas(ws):
    datos = seleccionarDatos(ws)
    db = crearConeccion()
    cursor = db.cursor()
    insStr = '''INSERT OR REPLACE INTO ig_personas (
                            nombres,
                            apellidos,
                            tipoIdentificacion,
                            identificacion,
                            codigoNomina,
                            email,
                            login,
                            password,
                            estado)
                        VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?)'''
    updStr = '''UPDATE ig_personas SET password = ? WHERE identificacion = ?'''
    cursor.executemany(insStr, datos)
    db.commit()
 
    #   Encriptar la contraseña (llave='palmeras1989')
    
    # for persona in datos:
    #     if persona[7] != '':
    #         identificacion = persona[3]
    #         passw = sha512.encrypt(persona[7])
    #         cursor.execute(updStr, (passw, identificacion))
    #         db.commit()
    db.close()

def carga_igPersonasLocales(ws):
    personas = crearIndiceIdentificacion(getPersonas())
    locales = crearIndiceCodLocal(getLocales())
    datos = seleccionarDatos(ws)
    valores=[(personas[dato[0]][0], locales[dato[1]][0], dato[2]) for dato in datos]
    db = crearConeccion()
    cursor = db.cursor()
    insStr = '''INSERT OR REPLACE INTO ig_personas_locales (
                            idPersona,
                            idLocal,
                            estado)
                        VALUES(?, ?, ?)'''
    cursor.executemany(insStr, valores)
    db.commit()
    db.close()

def carga_igPerfiles(ws):
    datos = seleccionarDatos(ws)
    db = crearConeccion()
    cursor = db.cursor()
    insStr = '''INSERT OR REPLACE INTO ig_perfiles (
                            perfil,
                            estado)
                        VALUES(?, ?)'''
    cursor.executemany(insStr, datos)
    db.commit()
    db.close()

def carga_igPerfilesPersonasLocal(ws):
    personas = crearIndiceIdentificacion(getPersonas())
    locales = crearIndiceCodLocal(getLocales())
    perfiles = crearIndicePerfil(getPerfiles())
    datos = seleccionarDatos(ws)
    # valores = []
    # for dato in datos:
    #     perfil = perfiles[dato[0]][0]
    #     persona = personas[dato[1]][0]
    #     local = locales[dato[2]][0]
    #     estado = dato[3]
    #     valores.append([perfil, persona, local, estado])
    valores=[(perfiles[dato[0]][0], personas[dato[1]][0], locales[dato[2]][0], dato[3]) for dato in datos]
    db = crearConeccion()
    cursor = db.cursor()
    insStr = '''INSERT OR REPLACE INTO ig_perfiles_personas_locales(
                            idPerfil,
                            idPersona,
                            idLocal,
                            estado)
                        VALUES(?, ?, ?, ?)'''
    cursor.executemany(insStr, valores)
    db.commit()
    db.close()

def carga_locGastosGenerales(ws):
    datos = seleccionarDatos(ws)
    db = crearConeccion()
    cursor = db.cursor()
    insStr = '''INSERT OR REPLACE INTO loc_gastosGenerales (
                            gastoGeneral,
                            cuentaContabilidad,
                            estado)
                        VALUES(?, ?, ?)'''
    cursor.executemany(insStr, datos)
    db.commit()
    db.close()

def carga_locGastosNoLocales(ws):
    locales = crearIndiceCodLocal(getLocales())
    gastosGenerales = crearIndiceGastosGenerales(getGastosGenerales())
    datos = seleccionarDatos(ws)
    valores = [(gastosGenerales[dato[0]][0], locales[dato[1]][0], dato[2]) for dato in datos]
    db = crearConeccion()
    cursor = db.cursor()
    insStr = '''INSERT OR REPLACE INTO loc_gastos_no_locales (
                            idgastoGeneral,
                            idLocal,
                            estado)
                        VALUES(?, ?, ?)'''
    cursor.executemany(insStr, valores)
    db.commit()
    db.close()

def carga_locFormasPagos(ws):
    datos = seleccionarDatos(ws)
    db = crearConeccion()
    cursor = db.cursor()
    insStr = '''INSERT OR REPLACE INTO loc_formasPagos (
                            formaPago,
                            cuentaContabilidad,
                            estado)
                        VALUES(?, ?, ?)'''
    cursor.executemany(insStr, datos)
    db.commit()
    db.close()

def carga_locCortesias(ws):
    datos = seleccionarDatos(ws)
    db = crearConeccion()
    cursor = db.cursor()
    insStr = '''INSERT OR REPLACE INTO loc_cortesias (
                            cortesia,
                            cuentaContabilidad,
                            estado)
                        VALUES(?, ?, ?)'''
    cursor.executemany(insStr, datos)
    db.commit()
    db.close()
    
def carga_locPagosPersonal(ws):
    datos = seleccionarDatos(ws)
    db = crearConeccion()
    cursor = db.cursor()
    insStr = '''INSERT OR REPLACE INTO loc_pagosPersonal (
                            tipoPago,
                            cuentaContabilidad,
                            estado)
                        VALUES(?, ?, ?)'''
    cursor.executemany(insStr, datos)
    db.commit()
    db.close()

def carga_baBancos(ws):
    datos = seleccionarDatos(ws)
    datos = seleccionarDatos(ws)
    db = crearConeccion()
    cursor = db.cursor()
    insStr = '''INSERT OR REPLACE INTO ba_bancos (
                            banco,
                            cuentaContabilidad,
                            estado)
                        VALUES(?, ?, ?)'''
    cursor.executemany(insStr, datos)
    db.commit()
    db.close()


def cargaDatosIniciales():
    datosInicialesFileName = os.getcwd() + '/src/CargaDatos/DatosIniciales.xlsx'
    wb = load_workbook(filename=datosInicialesFileName, data_only=True, read_only=True)
    for ws in wb:
        if ws.title == 'ig_locales':
            carga_igLocales(wb[ws.title])
        elif ws.title == 'ig_personas':
            carga_igPersonas(wb[ws.title])
        elif ws.title == 'ig_personasLocales':
            carga_igPersonasLocales(wb[ws.title])
        elif ws.title == 'ig_perfiles':
            carga_igPerfiles(wb[ws.title])
        elif ws.title == 'ig_perfilesPersonasLocales':
            carga_igPerfilesPersonasLocal(wb[ws.title])
        elif ws.title == 'loc_gastosGenerales':
            carga_locGastosGenerales(wb[ws.title])
        elif ws.title == 'loc_gastosNoLocales':
            carga_locGastosNoLocales(wb[ws.title])
        elif ws.title == 'loc_formasPagos':
            carga_locFormasPagos(wb[ws.title])
        elif ws.title == 'loc_cortesías':
            carga_locCortesias(wb[ws.title])
        elif ws.title == 'loc_pagosPersonal':
            carga_locPagosPersonal(wb[ws.title])
        elif ws.title == 'ba_bancos':
            carga_baBancos(wb[ws.title])

def crearDb():
    sqlitePath = os.getcwd() + '/data/PalmerasERP.db'
    if os.path.exists(sqlitePath):
        os.remove(sqlitePath)
    db = crearConeccion()
    db.close()

def crearTablas():
    scriptFile = os.getcwd() + '/src/modelo/crearTablasSQLite.sql'
    with open(scriptFile) as s:
        script = s.read()

    db = crearConeccion()
    cur = db.cursor()
    cur.executescript(script)
    db.commit()
    db.close()

def crearTriggers():
    db = crearConeccion()
    cur = db.cursor()
    triggersFileName = os.getcwd() + '/src/modelo/crearTriggersSQLite.txt'
    with open(triggersFileName) as triggerFile:
        triggers = triggerFile.read()
        for trigger in triggers.split(sep='$', maxsplit=-1):
            cur.executescript(trigger)
            db.commit()
    db.close()

def main():
    crearDb()
    crearTablas()
    crearTriggers()
    cargaDatosIniciales()

if __name__ == '__main__':
    main()